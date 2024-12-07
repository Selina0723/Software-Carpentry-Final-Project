import tkinter as tk
import tkinter.ttk
import tkinter.messagebox
from tkinter import messagebox
import sqlite3
from PIL import Image,ImageTk
import openpyxl
from tkinter import ttk, messagebox, filedialog
import json
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt


class App():
    def __init__(self):
        self.flag = True
        self.mainWindow = tk.Tk()
        self.mainWindow.title('Welcome to Rehabitation Task Manager!')
        self.mainWindow.geometry('450x450')
        
      


        self.canvas = tk.Canvas(self.mainWindow, height=450, width=450)
        self.imagefile = Image.open('bg.jpg')
        self.imagefile = ImageTk.PhotoImage(self.imagefile)
        self.image = self.canvas.create_image(0,0,anchor='nw',image=self.imagefile)
        self.canvas.pack(side='top')
        self.create_sqlite3()
        self.mainWindows()

    def create_sqlite3(self):
        database = "patient.db"
        self.conn = sqlite3.connect(database)
        self.cursor = self.conn.cursor()
        table_sql = '''
        Create Table if not exists userTable(
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            username varchar(30) not null,
            password varchar(20) not null
            );
            '''
        self.cursor.execute(table_sql)
        print("Database and user table created successfully!!!")

    def mainWindows(self):
        tk.Label(self.mainWindow, text='Username:').place(x=100, y=100)
        tk.Label(self.mainWindow, text='Password:').place(x=100, y=150)

        var_usr_name = tk.StringVar()
        self.entry_usr_name = tk.Entry(self.mainWindow,textvariable=var_usr_name)
        self.entry_usr_name.place(x=200, y=100)

        var_usr_pwd = tk.StringVar()
        self.entry_usr_pwd = tk.Entry(self.mainWindow, textvariable=var_usr_pwd, show='*')
        self.entry_usr_pwd.place(x=200, y=150)

        bt_login = tk.Button(self.mainWindow,text='LOG IN', bg='skyblue', command=self.usr_log_in)
        bt_login.place(x=100, y=250)
        bt_logup = tk.Button(self.mainWindow,text='CREATE', bg='skyblue', command=self.usr_sign_up)
        bt_logup.place(x=200, y=250)
        bt_logquit = tk.Button(self.mainWindow,text='QUIT', bg='skyblue', command=self.usr_sign_quit)
        bt_logquit.place(x=300, y=250)

    def usr_log_in(self):
        usr_name = self.entry_usr_name.get()
        usr_pwd = self.entry_usr_pwd.get()
        self.cursor.execute("SELECT * FROM userTable")
        rows = self.cursor.fetchall()

        if usr_name =='' or usr_pwd =='':
            messagebox.showerror(title='ERROR', message='Username or password is empty!')
        elif rows == []:
            is_signup = messagebox.askyesno('Welcome!',message='You have not registered yet, do you want to register now?')
            if is_signup:
                self.usr_sign_up()
        else:
            for row in rows:
                if usr_name == row[1]:
                    if usr_pwd == row[2]:
                        messagebox.showinfo(title='Welcome!', message='Welcome '+ usr_name)
                        self.usr_name = usr_name
                        self.mainWindow.destroy()
                        self.flag = False
                        func= Function(self)
                        func.mainWindows.mainloop()
                    else:
                        messagebox.showinfo(title='ERROR',message="WRONG PASSWORD")

    def usr_sign_up(self):
        def signtowcg():
            nn = self.usernameEntry.get()
            np = self.passwordEntry.get()
            npf = self.okPasswordEntry.get()
            self.cursor.execute('SELECT * FROM userTable')
            rows = self.cursor.fetchall()
            for row in rows:
                if nn in row:
                    tk.messagebox.showwarning('WARNING','Username already exists!')
                    break
            if np == '' or nn== '':
                tk.messagebox.showerror('ERROE','Username or password is empty!')
            elif np != npf:
                tk.messagebox.showerror('ERROR','The password is inconsistent!')
            
            else:
                insert_sql = """insert into userTable(username,password)values("{}", "{}");""".format(nn,np)
                self.cursor.execute(insert_sql)
                self.conn.commit()
                tk.messagebox.showinfo("Welcome", "Registered Successfully")
                self.registerWindows.destroy()

        self.registerWindows = tk.Tk()
        self.registerWindows.geometry('350x200')
        self.registerWindows.title('Register')
        #Username
        new_name = tk.StringVar()
        tk.Label(self.registerWindows, text='Username: ').place(x=10, y= 10)
        self.usernameEntry = tk.Entry(self.registerWindows, textvariable=new_name)
        self.usernameEntry.place(x=150, y=10)
        #Password
        new_pwd = tk.StringVar()
        tk.Label(self.registerWindows, text='Please enter your password:').place(x=10, y=50)
        self.passwordEntry = tk.Entry(self.registerWindows, textvariable=new_pwd, show='*')
        self.passwordEntry.place(x=150, y=50)
        #Enter again
        new_pwd_confirm = tk.StringVar()
        tk.Label(self.registerWindows, text='Please enter your password again:').place(x=10, y=90)
        self.okPasswordEntry = tk.Entry(self.registerWindows, textvariable=new_pwd_confirm, show='*')
        self.okPasswordEntry.place(x=150, y=90)
        # Confirm
        bt_confirm_sign_up = tk.Button(self.registerWindows, bg='red',text='Confirm',command=signtowcg)
        bt_confirm_sign_up.place(x=150, y=130)

    def usr_sign_quit(self):
        self.mainWindow.destroy()

    def run(self):
        self.mainWindow.mainloop()

class Function():
    def __init__(self,app_instance):
        self.app_instance = app_instance
        self.mainWindows = tk.Tk()
        self.mainWindows.resizable(False,False)
        self.mainWindows.title('Function Menu')
        self.mainWindows.geometry("500x700")
        # self.mainWindows = mainWindows
        self.tasks = []
        
    
    
        # self.load_tasks()
        self.linkSqlite3()
        self.interface()

    def linkSqlite3(self):
        database = "patient.db"
        self.conn = sqlite3.connect(database)
        self.cursor = self.conn.cursor()
        tabel_sql_patient = """
        Create Table if not exists patInfoTable(
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            patId varchar(20) not null,
            name varchar(20) not null,
            type varchar(20) not null,
            recovery varchar(20) not null
        );
        """

        # tabel_sql_task = """
        # CREATE TABLE IF NOT EXISTS taskInfoTable (
        #     taskID INTEGER PRIMARY KEY AUTOINCREMENT,
        #     patId VARCHAR(20) NOT NULL,
        #     taskNumber INTEGER NOT NULL,
        #     taskName NOT NULL,
        #     difficulty VARCHAR(20) NOT NULL,
        #     FOREIGN KEY (patId) REFERENCES patInfoTable(patId) ON DELETE CASCADE
        # );
        # """
        self.cursor.execute(tabel_sql_patient)
        #self.cursor.execute(tabel_sql_task)
        #self.conn.commit()
        print("The medical record table is created successfully!!!")

    def interface(self):
        bg_image = Image.open("bg4.jpg")  
        bg_image_resized = bg_image.resize((500,700))  
        self.bg_img = ImageTk.PhotoImage(bg_image_resized)
        bg_label = tk.Label(self.mainWindows, image=self.bg_img)
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)


        image = Image.open("image.jpg")
        resized_image = image.resize((60, 60))  
        self.img = ImageTk.PhotoImage(resized_image)  
        labelImage = tk.Label(self.mainWindows, image=self.img)
        labelImage.place(x=90, y=50) 
        #name
        usr_name = self.app_instance.usr_name
        BtuInsertStu = tk.Label(self.mainWindows, text="Hi " + usr_name + "!", font=("Arial -30 bold"))
        BtuInsertStu.place(x=230, y=50, height=60, width=200)
        
       
        # Add
        BtuInsertStu = tk.Button(self.mainWindows, text="Add Patient Information", font=("Arial -20 bold"), bg="lightgray",command=self.AddPatInfo)
        BtuInsertStu.place(x=90, y=150, height=40, width=300)
        # Delete
        BtuDeleteStu = tk.Button(self.mainWindows, text="Delete Patient Information", font=("Arial -20 bold"), bg="lightgray",command=self.DelPatInfo)
        BtuDeleteStu.place(x=90, y=200, height=40, width=300)
        # Edit
        BtuChangeStu = tk.Button(self.mainWindows, text="Edit Patient Information", font=("Arial -20 bold"), bg="lightgray",command=self.changePatInfo)
        BtuChangeStu.place(x=90, y=250, height=40, width=300)
        # Search
        BtuSelectStu = tk.Button(self.mainWindows, text="Search Patient Information", font=("Arial -20 bold"), bg="lightgray",command=self.selectPatInfo)
        BtuSelectStu.place(x=90, y=300, height=40, width=300)
        # Import
        BtuSelectStu = tk.Button(self.mainWindows, text="Import Patient Information", font=("Arial -20 bold"), bg="lightgray",command=self.importExcel)
        BtuSelectStu.place(x=90, y=350, height=40, width=300)
        # Export
        BtuSelectStu = tk.Button(self.mainWindows, text="Export Patient Information", font=("Arial -20 bold"), bg="lightgray",command=self.exportExcel)
        BtuSelectStu.place(x=90, y=400, height=40, width=300)
        #Task Management
        BtuSelectStu = tk.Button(self.mainWindows, text="Task\nManagement", font=("Arial -18 bold"), bg="lightgray",command=self.taskManagement)
        BtuSelectStu.place(x=70, y=500, height=120, width=120)
        #Progress
        BtuSelectStu = tk.Button(self.mainWindows, text="Task\nProgress", font=("Arial -18 bold"), bg="lightgray",command=self.taskProgress)
        BtuSelectStu.place(x=310, y=500, height=120, width=120)

    def AddPatInfo(self):
        self.addWindows = tk.Tk()
        self.addWindows.title("Add Patient Information")
        self.addWindows.geometry("500x450") 
        
        self.patIdVar = tk.StringVar(self.addWindows, value="")  
        self.nameVar = tk.StringVar(self.addWindows, value="")  
        self.TypeOfDiseaseVar = tk.StringVar(self.addWindows, value="")
        self.RecoveryTimeVar = tk.StringVar(self.addWindows, value="") 
        
        # Label
        patIdLabel = tk.Label(self.addWindows, text="Number:", font=("Arial -20"))
        patIdLabel.place(x=20, y=60, height=40, width=200)
        nameLabel = tk.Label(self.addWindows, text="Name:", font=("Arial -20"))
        nameLabel.place(x=20, y=110, height=40, width=200)
        TypeOfDiseaseLabel = tk.Label(self.addWindows, text="Type of Disease:", font=("Arial -20"))
        TypeOfDiseaseLabel.place(x=20, y=160, height=40, width=200)
        RecoveryTimeLabel = tk.Label(self.addWindows, text="Recovery Time(Days):", font=("Arial -20"))
        RecoveryTimeLabel.place(x=20, y=210, height=40, width=200)
        

        # Entry
        self.patIdEntry = tk.Entry(self.addWindows, textvariable=self.patIdVar)
        self.patIdEntry.place(x=230, y=60, height=40, width=200)
        self.nameEntry = tk.Entry(self.addWindows, textvariable=self.nameVar)
        self.nameEntry.place(x=230, y=110, height=40, width=200)
        self.TypeOfDiseaseEntry = tk.Entry(self.addWindows, textvariable=self.TypeOfDiseaseVar)
        self.TypeOfDiseaseEntry.place(x=230, y=160, height=40, width=200)
        self.RecoveryTimeEntry = tk.Entry(self.addWindows, textvariable=self.RecoveryTimeVar)
        self.RecoveryTimeEntry.place(x=230, y=210, height=40, width=200)
       

        def OkInsert():
            patId = self.patIdEntry.get()
            name = self.nameEntry.get()
            type = self.TypeOfDiseaseEntry.get()
            recovery = self.RecoveryTimeEntry.get()

           
            if not all([patId, name, type, recovery]):
                messagebox.showerror("ERROR", message="All fields are required!")
                return

            
            try:
                recovery = float(recovery)
            except ValueError:
                messagebox.showerror("ERROR", message="Recovery Time must be a valid number!")
                return

            
            self.cursor.execute("SELECT * FROM patInfoTable WHERE patId = ?", (patId,))
            num = self.cursor.fetchall()
            if num:
                messagebox.showwarning("WARNING", message="The patient ID already exists, please enter again!")
            else:
                try:
                    
                    sql1 = "INSERT INTO patInfoTable(patId, name, type, recovery) VALUES (?, ?, ?, ?)"
                    self.cursor.execute(sql1, (patId, name, type, recovery))
                    self.conn.commit()
                    messagebox.showinfo(title="Congratulations", message=f"{name}'s record has been successfully added!")
                except sqlite3.Error as e:
                    messagebox.showerror("ERROR", message=f"Failed to add record: {e}")


        # Add 
        InsertButton = tk.Button(self.addWindows, text="Add", font=("Arial -20"), bg="red",command=OkInsert)
        InsertButton.place(x=50, y=350, height=40, width=100)

        def Cancel():
            self.patIdVar.set("")
            self.nameVar.set("")
            self.TypeOfDiseaseVar.set("")
            self.RecoveryTimeVar.set("")
            

        # Reset
        CancelButton = tk.Button(self.addWindows, text="Reset", font=("Arial -20"), bg="green", command=Cancel)
        CancelButton.place(x=200, y=350, height=40, width=100)
        # Exit
        DeleteButton = tk.Button(self.addWindows, text="Exit", font=("Arial -20"), bg="blue", command=self.addWindows.destroy)
        DeleteButton.place(x=350, y=350, height=40, width=100)
        self.addWindows.mainloop()

    def DelPatInfo(self):
        self.delWindows = tk.Tk()
        self.delWindows.title("Delete Patient Information")
        self.delWindows.geometry("500x450")  
        # Label
        patIdLabel = tk.Label(self.delWindows, text="Number：", font=("Arial -20"))
        patIdLabel.place(x=20, y=100, height=40, width=200)
        self.patIdEntry = tk.Entry(self.delWindows)  
        self.patIdEntry.place(x=230, y=100, height=40, width=200)

        def OkDelete():  
            patId = self.patIdEntry.get()  
            if patId == "":
                messagebox.showerror("ERROR", message="Please Enter the Number!")
                return
            # Search
            self.cursor.execute("select * from patInfoTable where patId = '%s';" % patId)
            num = self.cursor.fetchall()
            if num:  
                self.cursor.execute("delete from patInfoTable where patId = '%s';" % patId)
                self.conn.commit()  
                messagebox.showinfo(title="Congratulations!", message="Successfully Delete！")
            else:  
                messagebox.showwarning("WARNING", message="The patient's information does not exist, please enter it first!")

        # Delete
        DeleteButton1 = tk.Button(self.delWindows, text="Delete", font=("Arial -20"), bg="red", command=OkDelete)
        DeleteButton1.place(x=150, y=220, height=40, width=200)

        # Exit
        DeleteButton2 = tk.Button(self.delWindows, text="Exit", font=("Arial -20"), bg="blue", command=self.delWindows.destroy)
        DeleteButton2.place(x=150, y=300, height=40, width=200)
        self.delWindows.mainloop()

    def changePatInfo(self):
        self.changeWindows = tk.Tk()
        self.changeWindows.title("Edit Patient Information")
        self.changeWindows.geometry("600x500")  
        
        self.changePatIdVar = tk.StringVar(self.changeWindows, value='')
        # Label
        changePatIdLabel = tk.Label(self.changeWindows, text="Number:", font=("Arial -20"))
        changePatIdLabel.place(x=130, y=10, height=40, width=100)
        self.changePatIdEntry = tk.Entry(self.changeWindows, textvariable=self.changePatIdVar)
        self.changePatIdEntry.place(x=250, y=10, height=40, width=200)

        def OkSelect():
            if self.changePatIdEntry.get() == "":
                messagebox.showerror(title="ERROR", message="Please Enter the Number!")
                return
            
            self.nameVar = tk.StringVar(self.changeWindows, value="")  
            self.TypeOfDiseaseVar = tk.StringVar(self.changeWindows, value="")  
            self.RecoveryTimeVar = tk.StringVar(self.changeWindows, value="")  
            self.patId = self.changePatIdEntry.get()  

            self.cursor.execute("select * from patInfoTable where patId = '%s';" % self.patId)
            l_row = self.cursor.fetchall()
            if l_row:
                for row in l_row:
                   
                    if self.patId == row[1]:
                        name = row[2]
                        TypeOfDisease = row[3]
                        RecoveryTime = row[4]
                        
                
                self.nameVar.set(name)
                self.TypeOfDiseaseVar.set(TypeOfDisease)
                self.RecoveryTimeVar.set(RecoveryTime)
                
            else:  
                messagebox.showwarning(title="WARNING", message="The patient information does not exist, please enter first!")

            patIdLabel = tk.Label(self.changeWindows, text="Number:", font=("Arial -20"))
            patIdLabel.place(x=100, y=160, height=40, width=200)
            nameLabel = tk.Label(self.changeWindows, text="Name:", font=("Arial -20"))
            nameLabel.place(x=100, y=210, height=40, width=200)
            TypeOfDiseaseLabel = tk.Label(self.changeWindows, text="Type of Disease:", font=("Arial -20"))
            TypeOfDiseaseLabel.place(x=100, y=260, height=40, width=200)
            RecoveryTimeLabel = tk.Label(self.changeWindows, text="Recovery Time(Days):", font=("Arial -20"))
            RecoveryTimeLabel.place(x=100, y=310, height=40, width=200)
            
    
            self.patIdLabel = tk.Label(self.changeWindows,text=self.patId)
            self.patIdLabel.place(x=320, y=160, height=40, width=95)
            self.nameEntry = tk.Entry(self.changeWindows, textvariable=self.nameVar)
            self.nameEntry.place(x=320, y=210, height=40, width=180)
            self.TypeOfDiseaseEntry = tk.Entry(self.changeWindows, textvariable=self.TypeOfDiseaseVar)
            self.TypeOfDiseaseEntry.place(x=320, y=260, height=40, width=180)
            self.RecoveryTimeEntry = tk.Entry(self.changeWindows, textvariable=self.RecoveryTimeVar)
            self.RecoveryTimeEntry.place(x=320, y=310, height=40, width=180)
           

            def saveInfo():  
                name = self.nameEntry.get()  
                TypeOfDisease = self.TypeOfDiseaseEntry.get()  
                
            
                
                if name == "" or TypeOfDisease == "" or self.RecoveryTimeEntry.get() == "" :
                    messagebox.showerror("ERROR", message="The message cannot be empty!")
                    return
                RecoveryTime = eval(self. RecoveryTimeEntry.get())
                if RecoveryTime<0 or RecoveryTime>100000000:
                    messagebox.showerror("ERROR", message="Patient information entered incorrectly!!!")
                else:
                    UpdateSql = "update patInfoTable set name='%s',type='%s',recovery=%f where patId='%s';" \
                                    % (name,TypeOfDisease,RecoveryTime,self.patId)
                    self.cursor.execute(UpdateSql)
                    self.conn.commit()
                    messagebox.showinfo(title='Congratulations', message='Patient information edited successfully!！')
            # Edit
            saveButton = tk.Button(self.changeWindows, text="Edit", font=("Arial -20"), bg="lightgray", command=saveInfo)
            saveButton.place(x=200, y=430, height=40, width=200)

        # Select
        SelectButton = tk.Button(self.changeWindows, text="Select", font=("Arial -20"), bg="red", command=OkSelect)
        SelectButton.place(x=100, y=80, height=40, width=100)

        def Cancel():
            self.changePatIdVar.set('')
        CancelButton = tk.Button(self.changeWindows, text="Cancel", font="Arial -20", bg="green", command=Cancel)
        CancelButton.place(x=250, y=80, height=40, width=100)
        # Exit
        DeleteButton = tk.Button(self.changeWindows, text="Exit", font="Arial -20", bg="blue", command=self.changeWindows.destroy)
        DeleteButton.place(x=400, y=80, height=40, width=100)

        self.changeWindows.mainloop()

    def selectPatInfo(self):
        self.selectWindows = tk.Tk()
        self.selectWindows.title("Search Patient Information")
        self.selectWindows.geometry("600x600")  # 大小
        self.selectWindows_frame1 = tk.Frame(self.selectWindows,width=600,height=100)
        self.selectWindows_frame1.place(x=0,y=0)
        self.selectWindows_frame2 = tk.Frame(self.selectWindows, width=600, height=500)
        self.selectWindows_frame2.place(x=0, y=100)
        
        self.selectPatIdVar = tk.StringVar(self.selectWindows_frame1, value='')
        
        selectPatIdLabel = tk.Label(self.selectWindows_frame1, text="Number:", font=("Arial -20"))
        selectPatIdLabel.place(x=10, y=30, height=40, width=80)
        self.selectPatIdEntry = tk.Entry(self.selectWindows_frame1, textvariable=self.selectPatIdVar)
        self.selectPatIdEntry.place(x=100, y=30, height=40, width=150)

        def OkSelect():
            self.patId = self.selectPatIdEntry.get()  
            if self.patId == "":
                messagebox.showerror("ERROR", message="The patient number cannot be empty！")
                return
            
            self.cursor.execute("select * from patInfoTable where patId = '%s';" % self.patId)
            l_row = self.cursor.fetchall()
            if l_row:
                for row in l_row:
                   
                    if self.patId == row[1]:
                        name = row[2]
                        TypeOfDisease = row[3]
                        RecoveryTime = row[4]
                       
                        tree = tkinter.ttk.Treeview(self.selectWindows_frame2, columns=('c1', 'c2', 'c3', 'c4'),
                                                    show='headings', height=500)
                        tree.column('c1', width=95, anchor='center')
                        tree.column('c2', width=65, anchor='center')
                        tree.column('c3', width=70, anchor='center')
                        tree.column('c4', width=70, anchor='center')
                        
                        tree.heading('c1', text="Number")
                        tree.heading('c2', text="Name")
                        tree.heading('c3', text="Type of Disease")
                        tree.heading('c4', text="Recovery Time")
                     
                        # Treeview
                        tree.place(relx=0.0, rely=0.0, relwidth=1.0, relheight=1.0)
                        ls_data = [[self.patId, name, TypeOfDisease, RecoveryTime]]
                        for data in ls_data:
                            tree.insert('', 'end', text='', values=data)
                        break
            else:
                messagebox.showwarning("WARNING", message="The patient information does not exist, please enter it first!")
        #search all
        def OkSelectAll():
            
            self.cursor.execute("select * from patInfoTable;")
            l_row = self.cursor.fetchall()
            l_data = []
            for row in l_row:
                patId = row[1]
                name = row[2]
                TypeOfDisease = row[3]
                RecoveryTime = row[4]
               
                data = [patId, name, TypeOfDisease, RecoveryTime]
                l_data.append(data)

            tree = tkinter.ttk.Treeview(self.selectWindows_frame2,
                                        columns=('c1', 'c2', 'c3', 'c4'),
                                        show='headings', height=500)
            scrollbar = tkinter.Scrollbar(self.selectWindows_frame2, bg="blue",command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            tree.column('c1', width=95, anchor='center')
            tree.column('c2', width=65, anchor='center')
            tree.column('c3', width=70, anchor='center')
            tree.column('c4', width=70, anchor='center')
            
            tree.heading('c1', text="Number")
            tree.heading('c2', text="Name")
            tree.heading('c3', text="Type of Disease")
            tree.heading('c4', text="Recovery Time")
            
            # Treeview
            scrollbar.place(relx=1.0, rely=0.0, relheight=1.0, anchor='ne')
            tree.place(relx=0.0, rely=0.0, relwidth=1.0, relheight=1.0)
            for data in l_data:
                tree.insert('', 'end', text='', values=data)
        # S
        SelectButton = tk.Button(self.selectWindows_frame1, text="Select", font=("Arial -20"), bg="red", command=OkSelect)
        SelectButton.place(x=270, y=30, height=40, width=70)
        # 
        SelectAllButton = tk.Button(self.selectWindows_frame1, text="Search All", font=("Arial -20"), bg="red", command=OkSelectAll)
        SelectAllButton.place(x=360, y=30, height=40, width=130)
        # 
        DeleteButton = tk.Button(self.selectWindows_frame1, text="Exit", font="Arial -20", bg="blue", command=self.selectWindows.destroy)
        DeleteButton.place(x=510, y=30, height=40, width=70)
        self.selectWindows.mainloop()

    def importExcel(self):
        file_path = filedialog.askopenfilename(title="Open The File",filetypes = [("Excel File","xlsx")])
        if file_path == "":
            messagebox.showwarning("WARNING", message="Please re-import!!")
        else:
            try:
                # Excel
                workbook = openpyxl.load_workbook(file_path)
                
                sheet = workbook.active
                all_data = []
              
                for row in sheet.iter_rows():
                    row_data = []
                    for cell in row:
                        row_data.append(cell.value)
                    all_data.append(row_data)
                    
                self.cursor.execute("select * from patInfoTable;")
                l_row = self.cursor.fetchall()
                for data in all_data[1:]:
                    patId = data[1]
                    self.cursor.execute("select * from patInfoTable where stuId = '%s';" % patId)
                    l_row = self.cursor.fetchall()
                    if l_row:
                        continue
                    name = data[2]
                    TypeOfDisease = data[3]
                    RecoveryTime = float(data[4])
              
                    try:  #
                        sql1 = "INSERT INTO patInfoTable(patId,name,TypeOfDisease, RecoveryTime)"
                        sql1 += "VALUES('%s','%s','%s',%f)" % (patId,name,TypeOfDisease, RecoveryTime)
                        self.cursor.execute(sql1)  
                        self.conn.commit()  
                    except:
                        messagebox.showerror("WARNING", message = "The format of patient results information is wrong and cannot be entered!")
                messagebox.showwarning(title="WARNING", message="If the patient number is repeated, it cannot be entered!！")
                messagebox.showinfo(title="Congratulations!", message="Patient information batch entry success!")
            except:
                messagebox.showerror("ERROR", message="The form format is wrong, please re-select!!")

    def exportExcel(self):
        file_path = filedialog.asksaveasfilename(title="Export File", initialfile="pat_data.xlsx",
                                             filetypes=[("Excel 工作簿", "*.xlsx")])
        if file_path == "":
            messagebox.showwarning("WARNING", message="Please re-export！")
        else:
            workbook = openpyxl.Workbook()  
            sheet = workbook.active  
            sheet["A1"] = "Sequence"
            sheet["B1"] = "Number"
            sheet["C1"] = "Name"
            sheet["D1"] = "Type Of Disease"
            sheet["E1"] = "Recovery Time"
        
            self.cursor.execute("select * from patInfoTable;")
            rows = self.cursor.fetchall()
            for row in rows:
                sheet.append(row)
            workbook.save(file_path)
            messagebox.showinfo(title="Congratulations", message="Patient information exported successfully!")

 
   

    def taskManagement(self):
        self.taskWindows = tk.Tk()
        self.taskWindows.title("Task Management")
        self.taskWindows.geometry("500x450")
        # self.tasks= []

        # Task Name
        tk.Label(self.taskWindows, text="Task:").place(x=10, y=10, width=50, height=25)
        self.task_name_entry = tk.Entry(self.taskWindows, width=30)
        self.task_name_entry.place(x=70, y=10, width=200, height=25)

        # Difficulty
        tk.Label(self.taskWindows, text="Difficulty:").place(x=10, y=50, width=70, height=25)
        self.difficulty_var = tk.StringVar(value="Easy")
        ttk.Combobox(self.taskWindows, textvariable=self.difficulty_var, values=["Easy", "Medium", "Hard"],state="readonly").place(x=90, y=50, width=150, height=25)
     

        # Task List
        self.task_listbox = tk.Listbox(self.taskWindows, height=10, selectmode="single")
        self.task_listbox.place(x=10, y=90, width=460, height=150)

    

        # Force Refresh
        self.taskWindows.update()
        #self.update_task_list()
        # Sub-functions
        def add_task():
            task_name = self.task_name_entry.get().strip()
            difficulty = self.difficulty_var.get()
            if task_name:
                self.tasks.append({"task": task_name, "difficulty": difficulty, "completed": False})
                update_task_list()
                self.task_name_entry.delete(0, tk.END)
            else:
                messagebox.showerror("Error", "Task name cannot be empty!")

        

        def update_task_list():
            self.task_listbox.delete(0, tk.END)
            for task in self.tasks:
                status = "[Done]" if task["completed"] else "[Pending]"
                self.task_listbox.insert(tk.END, f"{status} {task['task']} ({task['difficulty']})")

        def mark_complete():
            selected = self.task_listbox.curselection()
            if selected:
                index = selected[0]
                self.tasks[index]["completed"] = True
                update_task_list()
            else:
                messagebox.showwarning("Warning", "No task selected!")

        def delete_task():
            selected = self.task_listbox.curselection()
            if selected:
                index = selected[0]
                del self.tasks[index]
                update_task_list()
            else:
                messagebox.showwarning("Warning", "No task selected!")

        def save_tasks():
            file_path = filedialog.asksaveasfilename(
                title="Save Tasks", defaultextension=".json", filetypes=[("JSON files", "*.json")]
            )
            if file_path:
                with open(file_path, "w") as f:
                    json.dump(self.tasks, f)
                messagebox.showinfo("Success", "Tasks saved successfully!")

        def load_tasks():
            try:
                with open("tasks.json", "r") as f:
                    self.tasks = json.load(f)
            except FileNotFoundError:
                self.tasks = []

        tk.Button(self.taskWindows, text="Add Task", command=add_task).place(x=300, y=10, width=100, height=25)
        tk.Button(self.taskWindows, text="Mark as Complete", command=mark_complete).place(x=10, y=250, width=150, height=25)
        tk.Button(self.taskWindows, text="Delete Task", command=delete_task).place(x=170, y=250, width=150, height=25)
        tk.Button(self.taskWindows, text="Save Tasks", command=save_tasks).place(x=330, y=250, width=150, height=25)
        tk.Button(self.taskWindows, text="Exit", command=self.taskWindows.destroy).place(x=170, y=320, width=150, height=25)

  
    def taskProgress(self):
        self.progressWindows = tk.Toplevel(self.mainWindows)
        self.progressWindows.title("Progress Visualization")
        self.progressWindows.geometry("500x450")

        # Progress visualization frame
        progress_frame = ttk.LabelFrame(self.progressWindows, text="Progress Visualization")
        progress_frame.place(x=10, y=10, width=480, height=350)

        # Progress Plot
        self.figure, self.ax = plt.subplots(figsize=(5, 3))  
        self.canvas = FigureCanvasTkAgg(self.figure, master=progress_frame) 
        self.canvas.get_tk_widget().place(x=10, y=10, width=460, height=300)

        # Exit Button
        tk.Button(self.progressWindows, text="Exit", command=self.progressWindows.destroy).place(x=200, y=380, width=100, height=25)

        
        self.plot_progress()

    def plot_progress(self):
        self.ax.clear()  

        
        total_tasks = len(self.tasks)
        completed_tasks = sum(1 for task in self.tasks if task["completed"])
        pending_tasks = total_tasks - completed_tasks

        if total_tasks > 0:
            self.ax.pie(
                [completed_tasks, pending_tasks],
                labels=["Completed", "Pending"],
                autopct="%1.1f%%",
                colors=["#4CAF50", "#FFC107"],  
            )
        else:
    
            self.ax.text(0.5, 0.5, "No Tasks", horizontalalignment="center", verticalalignment="center", fontsize=12)

        self.ax.set_title("Task Completion Progress")  
        self.canvas.draw() 
    
    def run(self):
        self.mainWindows.mainloop()

app = App()
app.run()
if app.flag == False:
    fun = Function()
    fun.run()